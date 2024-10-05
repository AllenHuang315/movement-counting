#-*- coding:utf-8 -*-
import openpyxl as pyxl
import numpy as np
import math
import os
from collections import OrderedDict


def parseXLSX(xlsxName, parseArray):
    wb = pyxl.load_workbook(xlsxName)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    for i in range(2, ws.max_row+1):
        cvalue = ws.cell(row = i, column = 2).value # frame_id
        if cvalue is not None:
            fId = int(cvalue)
        else:
            import pdb; pdb.set_trace() #debug use
        cvalue = ws.cell(row = i, column = 3).value # movement_id
        if cvalue is not None:
            mId = int(cvalue)
        else:
            import pdb; pdb.set_trace()
        cvalue = ws.cell(row = i, column = 4).value # class_id
        if cvalue is not None:
            if cvalue == 0:
                type = 2
            elif cvalue == 1:
                type = 4
            elif cvalue == 2:
                type = 3
            elif cvalue == 3:
                type = 9
            elif cvalue == 4:
                type = 5
            elif cvalue == 5:
                type = 8
            elif cvalue == 6:
                type = 1
            elif cvalue == 7:
                type = 11
            elif cvalue == 8:
                type = 6
            elif cvalue == 9:
                type = 7
            elif cvalue == 10:
                type = 10
            elif cvalue == 11:
                type = 12
            elif cvalue == 12:
                type = 13 
        else:
            import pdb; pdb.set_trace()
    
        parseArray[fId, mId - 1, type] = 1 # [frame_Id, movement_Id, class_Id]


def parseTxt(TxtName, parseArray):
    fp = open(TxtName)
    lines = fp.readlines()
    for line in lines:
        sp = line.rstrip("\n").split(",")
        fId = int(sp[1])
        #if fId >= 3000:
        #    continue
        mId = int(sp[2])
        cvalue = int(sp[3]) # cvalue 0- 12 對應 class_id 1-13
        if cvalue == 0:
            type = 2
        elif cvalue == 1:
            type = 4
        elif cvalue == 2:
            type = 3
        elif cvalue == 3:
            type = 9
        elif cvalue == 4:
            type = 5
        elif cvalue == 5:
            type = 8
        elif cvalue == 6:
            type = 1
        elif cvalue == 7:
            type = 11
        elif cvalue == 8:
            type = 6
        elif cvalue == 9:
            type = 7
        elif cvalue == 10:
            type = 10
        elif cvalue == 11:
            type = 12
        elif cvalue == 12:
            type = 13 

        parseArray[fId, mId - 1, type] = 1 # (frame_id, movement_id, class_id)

def compute_nwRMSE(n, pdArray, gtArray):
    # weight
    wVect = np.asarray(np.arange(1, n+1)) / (n * (n + 1) / 2.0) #???
    fNum, mNum, typeNum = pdArray.shape #[123, 17, 2]
    lst = range(0, fNum) # range(0,123)
    interval = int(math.ceil(fNum / float(n))) # 13
    segLsts = [lst[i : i + interval] for i in range(0, len(lst), interval)] #0-13 14-26...
    gtCntArray = np.zeros(mNum) # 1d
    pdCntArray = np.zeros(mNum)
    nwRMSEArray = np.zeros((mNum, 13)) # 2d
    wRMSEArray = np.zeros((mNum, 13))
    vehicleNumArray = np.zeros((mNum, 13))
    for mId in range(0, mNum):
        gtCntArray[mId] = np.sum(gtArray[:, mId, :]) # CLASS_ID, TRAJECTORY_ID MISTAKE COUNTING SUM 
        pdCntArray[mId] = np.sum(pdArray[:, mId, :])
        for tId in range(0, 13):
            # wRMSE
            diffVectCul = np.zeros(n)
            for segId, frames in enumerate(segLsts):
                diff = np.square(sum(pdArray[0:frames[-1], mId, tId]) - sum(gtArray[0:frames[-1], mId, tId]))
                diffVectCul[segId] = diff
            wRMSE = np.sqrt(np.dot(wVect, diffVectCul))

            # num
            vehicleNum = np.sum(gtArray[:, mId, tId]) # 指定movement_id跟class_id 有多少vehicles
            vehicleNumArray[mId, tId] = vehicleNum 

            # for print only
            if vehicleNum == 0:
                wRMSEArray[mId, tId] = 0 
            else:
                wRMSEArray[mId, tId] = wRMSE / vehicleNum

            #nwRMSE
            if wRMSE > vehicleNum:
                nwRMSE = 0
            else:
                if vehicleNum == 0:
                    nwRMSE = 0
                else:
                    nwRMSE = 1 - wRMSE / vehicleNum
            nwRMSEArray[mId, tId] = nwRMSE

    print("")
    printStr = " moveID: "
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "% 4d | "%(moveId+1)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " gt cnt: "
    for val in gtCntArray.tolist():
        printStr += "%04d | "%(val)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " pd cnt: "
    for val in pdCntArray.tolist():
        printStr += "%04d | "%(val)
    print(printStr)
    printStr = " --------"
    for moveId, val in enumerate(np.sum(wRMSEArray, axis=1).tolist()):
        printStr += "-------"
    print(printStr)

    printStr = " nwRMSE: "
    for moveId, val in enumerate(np.sum(nwRMSEArray, axis=1).tolist()):
        printStr += "%2.2f | "%(val)
    print(printStr)

    nwRMSEArray = np.multiply(nwRMSEArray, vehicleNumArray)
    return np.sum(nwRMSEArray), np.sum(vehicleNumArray) 


if __name__ == "__main__":
    videoInfo = {"cam_1":{"frame_num":6000, "movement_num":20}}

    # segment number
    n = 10 

    #gtXlsxRoot = "./gt/"
    #pdXlsxRoot = "./vehicle_counting_results/"
    vNum = len(videoInfo.keys()) #
    nwRMSEVec = np.zeros(vNum) # 均為0 長度6000的 1D array 
    vehicleNumVec = np.zeros(vNum) 
    vId = 0
    for vName, info in videoInfo.items():
        fNum = videoInfo[vName]["frame_num"] # 6000
        #if fNum > 3000:
        #    fNum = 3000
        mNum = videoInfo[vName]["movement_num"] # 20

        # parse gt
        gtArray = np.zeros((fNum, mNum, 13)) #  (共幾組, row, col) # 13 classes
        gtXlsx = "C:/Users/Allen/Desktop/tracking/StrongSORT-YOLO/VehicleCounting/eval/gt/cam_1.xlsx"
        if not os.path.exists(gtXlsx):
            continue
        parseXLSX(gtXlsx, gtArray)

        # parse prediction 
        pdArray = np.zeros((fNum, mNum, 13))
        pdXlsx = "C:/Users/Allen/Desktop/tracking/StrongSORT-YOLO/VehicleCounting/eval/vehicle_counting_results/cam_1.txt"
        if not os.path.exists(pdXlsx):
            continue
        parseTxt(pdXlsx, pdArray)

        nwRMSE, vehicleNum = compute_nwRMSE(n, pdArray, gtArray)
        #nwRMSEVec[vId] = nwRMSE
        #vehicleNumVec[vId] = vehicleNum
        vId += 1
        print(" %s nwRMSE: %f"%(vName, nwRMSE/vehicleNum))

    score2 = sum(nwRMSEVec) / sum(vehicleNumVec)

    baseFactor = 0.569127 # from efficiency_base.py
    videoTotal = 300 # one 300 sec video
    #time = 6217 
    time = 838 # res50 #處理影片所花的時間 tracking #tucson_1 1487.466 #tucson_2 1599.859
    #time = 11418 # res50 pipeline
    #time = 43642 # res154
    #time = 8487 # omni 
    score1 = 1 - (time * baseFactor) / (5 * float(videoTotal)) 

    score = 0.3 * score1 + 0.7 * score2
    print("\ns1: %f; effective: %f; efficient: %f"%(score, score2, score1))
